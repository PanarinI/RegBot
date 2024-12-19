import os
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
import re

def main():
    # Проверяем наличие файла _register_test.xlsx в текущей папке
    file_name = "_register_test.xlsx"
    if not os.path.exists(file_name):
        print(f"Файл {file_name} не найден. Введите абсолютный путь к файлу:")
        while True:
            file_path = input("Введите путь к файлу: ").strip()
            if os.path.exists(file_path) and file_path.endswith((".xlsx", ".xlsm")):
                file_name = file_path
                break
            print("Файл не найден или это не Excel-файл. Попробуйте снова.")
    else:
        print(f"Файл {file_name} обнаружен.")

    # Проверка ввода ключевого слова "реестр"
    while True:
        word = input("Введите слово 'реестр': ").strip().lower()
        if word == "реестр":
            break
        print("Нет, именно 'реестр':)")

    # Главное меню
    while True:
        print("\nГлавное меню:")
        print("1. Открыть реестр в Excel")
        print("2. Добавить запись")
        print("3. Изменить запись")
        print("4. Настроить таблицу")
        print("5. Выход")

        choice = input("Выберите пункт меню: ").strip()

        if choice == "1":
            open_registry(file_name)
        elif choice == "2":
            add_record(file_name)
        elif choice == "3":
            edit_record(file_name)
        elif choice == "4":
            configure_table(file_name)
        elif choice == "5":
            print("Выход из программы. До свидания!")
            break
        else:
            print("Некорректный ввод. Выберите пункт от 1 до 5.")

# Открыть реестр в Excel
def open_registry(file_name):
    try:
        print(f"Открываю файл {file_name}...")
        os.startfile(file_name)
    except Exception as e:
        print(f"Не удалось открыть файл. Ошибка: {e}")

# Добавить запись
def add_record(file_name):
    try:
        workbook = load_workbook(file_name)
        sheet = workbook.active
    except Exception as e:
        print(f"Ошибка при открытии файла: {e}")
        return

    print("\nДобавление новой записи")
    first_empty_row = find_first_empty_row(sheet)

    # Заголовки столбцов
    headers = [sheet.cell(row=1, column=col).value for col in range(1, sheet.max_column + 1)]
    section_choices = [
        "Банк практик", "Видеолекции", "Видеоматериалы", "Жизненные ситуации",
        "Инструкция по внедрению", "Исследования и обзоры", "Публикации",
        "Отечественный опыт", "Международный опыт", "Иное"
    ]
    new_record = {}

    for col, header in enumerate(headers, start=1):
        if not header:
            continue

        col_letter = get_column_letter(col)
        while True:
            if col_letter == "A" or col_letter == "B":
                value = input(f"{header} (1 - 'Нет', 2 - 'Да'): ").strip()
                if value == "1":
                    value = "Нет"
                elif value == "2":
                    value = "Да"
                else:
                    print("Некорректный ввод. Введите 1 или 2.")
                    continue
                break
            elif col_letter == "F":
                print("Выберите раздел:")
                for i, choice in enumerate(section_choices, start=1):
                    print(f"{i}. {choice}")
                choice = input("Введите номер раздела: ").strip()
                if choice.isdigit() and 1 <= int(choice) <= len(section_choices):
                    value = section_choices[int(choice) - 1]
                    if value == "Иное":
                        value = input("Введите название для 'Иное': ").strip()
                    break
                print("Некорректный выбор. Введите номер из списка.")
            elif col_letter == "G":
                value = input(f"Введите значение для {header} (или оставьте пустым): ").strip()
                break
            elif col_letter == "H" or col_letter == "I":
                value = input(f"Введите значение для {header} (в формате ДД-ММ-ГГГГ или оставьте пустым): ").strip()
                if not value:  # Поле может быть пустым
                    break
                if re.match(r"^\d{2}-\d{2}-\d{4}$", value):
                    break
                print(f"Некорректный формат даты. Введите дату в формате ДД-ММ-ГГГГ или оставьте пустым.")
            elif col_letter == "J":
                value = input(f"{header} (1 - 'Нет', 2 - 'Да'): ").strip()
                if value == "1":
                    print("Тогда разместите файл и потом выберите 'Да'.")
                    continue
                elif value == "2":
                    value = "Да"
                else:
                    print("Некорректный ввод. Введите 1 или 2.")
                    continue
                break
            elif col_letter == "M":
                value = input(f"{header}: ").strip()
                if "<iframe src=" in value:
                    break
                print("Введите корректный код внедрения OneDrive.")
            elif col_letter == "N":
                iframe_code = new_record.get("Код внедрения (OneDrive, если текст)", "")
                if iframe_code:
                    adapted_code = re.sub(r'width="[^"]*"', 'width="90%"', iframe_code)
                    adapted_code = re.sub(r'height="[^"]*"', 'height="1800"', adapted_code)
                    value = f'<p align="center">{adapted_code}</p>'
                    print(f"Код внедрения адаптирован для отображения на tilda: {value}")
                else:
                    value = None
                break
            elif col_letter == "O":
                value = input(f"Введите имя страницы Tilda (латиница): ").strip()
                if value.isalnum():
                    break
                print("Поле может содержать только латинские буквы и цифры. Попробуйте снова.")
            elif col_letter == "P":
                value = input(f"{header} (или оставьте пустым): ").strip()
                break
            else:
                value = input(f"Введите значение для {header}: ").strip()
                if value:
                    break
                print(f"Поле {header} обязательно для заполнения. Попробуйте снова.")

        new_record[header] = value
        sheet.cell(row=first_empty_row, column=col).value = value

    workbook.save(file_name)
    print("Запись успешно добавлена.")

# Изменить запись
def edit_record(file_name):
    print("Изменение записи пока не реализовано.")

# Настроить таблицу
def configure_table(file_name):
    print("Настройка таблицы пока не реализована.")

# Найти первую пустую строку
def find_first_empty_row(sheet):
    for row in range(2, sheet.max_row + 2):
        if not any(sheet.cell(row=row, column=col).value for col in range(1, sheet.max_column + 1)):
            return row

if __name__ == "__main__":
    main()
