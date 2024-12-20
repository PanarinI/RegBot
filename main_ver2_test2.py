import os
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
import re

def main():
    # Проверяем наличие файла _register_test.xlsx в текущей папке
    file_name = "_register_test.xlsx"
    if not os.path.exists(file_name):
        print(f"Файл {file_name} не найден. Введите абсолютный путь к файлу:")
        while True:
            file_path = input("Введите путь к файлу: ").strip()
            if os.path.exists(file_path) and file_path.endswith(('.xlsx', '.xlsm')):
                file_name = file_path
                break
            print("Файл не найден или это не Excel-файл. Попробуйте снова.")
    else:
        print(f"Файл {file_name} обнаружен.")

    # Проверка ввода ключевого слова "реестр"
    while True:
        word = input("Введите слово 'реестр': ").strip().lower()
        if word == "реестр":
            break
        print("Нет, именно 'реестр':)")

    # Главное меню
    while True:
        print("\nГлавное меню:")
        print("1. Открыть реестр в Excel")
        print("2. Добавить запись")
        print("3. Изменить запись")
        print("4. Настроить таблицу")
        print("5. Выход")

        choice = input("Выберите пункт меню: ").strip()

        if choice == "1":
            open_registry(file_name)
        elif choice == "2":
            add_record(file_name)
        elif choice == "3":
            edit_record(file_name)
        elif choice == "4":
            configure_table(file_name)
        elif choice == "5":
            print("Выход из программы. До свидания!")
            break
        else:
            print("Некорректный ввод. Выберите пункт от 1 до 5.")

# Открыть реестр в Excel
def open_registry(file_name):
    try:
        print(f"Открываю файл {file_name}...")
        os.startfile(file_name)
    except Exception as e:
        print(f"Не удалось открыть файл. Ошибка: {e}")

# Универсальные функции проверки полей
def validate_yes_no(header):
    while True:
        value = input(f"{header} (1 - 'Нет', 2 - 'Да'): ").strip()
        if value == "1":
            return "Нет"
        elif value == "2":
            return "Да"
        else:
            print("Некорректный ввод. Введите 1 или 2.")

def validate_required_text(header):
    while True:
        value = input(f"Введите значение для {header}: ").strip()
        if value:
            return value
        print(f"Поле {header} обязательно для заполнения. Попробуйте снова.")

def validate_optional_text(header):
    return input(f"Введите значение для {header} (или оставьте пустым): ").strip()

def validate_section(header):
    section_choices = [
        "Банк практик", "Видеолекции", "Видеоматериалы", "Жизненные ситуации",
        "Инструкция по внедрению", "Исследования и обзоры", "Публикации",
        "Отечественный опыт", "Международный опыт", "Иное"
    ]
    while True:
        print("Выберите раздел:")
        for i, choice in enumerate(section_choices, start=1):
            print(f"{i}. {choice}")
        choice = input("Введите номер раздела: ").strip()
        if choice.isdigit() and 1 <= int(choice) <= len(section_choices):
            section = section_choices[int(choice) - 1]
            if section == "Иное":
                return input("Введите название для 'Иное': ").strip()
            return section
        print("Некорректный выбор. Попробуйте снова.")

def validate_date(header):
    while True:
        value = input(f"Введите значение для {header} (в формате ДД-ММ-ГГГГ или оставьте пустым): ").strip()
        if not value or re.match(r"^\d{2}-\d{2}-\d{4}$", value):
            return value
        print("Некорректный формат даты. Введите дату в формате ДД-ММ-ГГГГ или оставьте пустым.")

def validate_final_folder(header):
    while True:
        value = input(f"{header} (1 - 'Нет', 2 - 'Да'): ").strip()
        if value == "1":
            print("Тогда разместите файл и потом выберите 'Да'.")
            continue
        elif value == "2":
            return "Да"
        else:
            print("Некорректный ввод. Введите 1 или 2.")

def validate_iframe(header):
    while True:
        value = input(f"{header}: ").strip()
        if "<iframe src=" in value:
            return value
        print("Введите корректный код внедрения OneDrive.")

def validate_tilda_name(header):
    while True:
        value = input(f"{header} (латиница): ").strip()
        if value.isascii() and value.isalnum():
            return value
        print("Поле может содержать только латинские буквы и цифры. Попробуйте снова.")

def validate_format(header):
    while True:
        value = input(f"{header} (1 - 'Текст', 2 - 'Видео'): ").strip()
        if value == "1":
            return "Текст"
        elif value == "2":
            return "Видео"
        else:
            print("Некорректный ввод. Введите 1 или 2.")

def validate_field(header, col_letter, current_value=None):
    if col_letter in {"A", "B"}:
        return validate_yes_no(header)
    elif col_letter == "C" or col_letter == "D":
        return validate_required_text(header)
    elif col_letter == "E":
        return validate_required_text(header)
    elif col_letter == "F":
        return validate_section(header)
    elif col_letter == "G":
        return validate_optional_text(header)
    elif col_letter == "H" or col_letter == "I":
        return validate_date(header)
    elif col_letter == "J":
        return validate_final_folder(header)
    elif col_letter == "K":
        return validate_required_text(header)
    elif col_letter == "L":
        return validate_format(header)
    elif col_letter == "M":
        if current_value == "Видео":
            print("Формат 'Видео'. Поле пропускается.")
            return None
        return validate_iframe(header)
    elif col_letter == "N":
        if current_value == "Видео":
            print("Формат 'Видео'. Поле пропускается.")
            return None
        iframe_code = current_value
        adapted_code = re.sub(r'width="[^"]*"', 'width="90%"', iframe_code)
        adapted_code = re.sub(r'height="[^"]*"', 'height="1800"', adapted_code)
        return f'<p align="center">{adapted_code}</p>'
    elif col_letter == "O":
        return validate_tilda_name(header)
    elif col_letter == "P":
        return validate_optional_text(header)

# Добавить запись
def add_record(file_name):
    try:
        workbook = load_workbook(file_name)
        sheet = workbook.active
    except Exception as e:
        print(f"Ошибка при открытии файла: {e}")
        return

    print("\nДобавление новой записи")
    first_empty_row = find_first_empty_row(sheet)

    headers = [sheet.cell(row=1, column=col).value for col in range(1, sheet.max_column + 1)]
    new_record = {}

    for col, header in enumerate(headers, start=1):
        if not header:
            continue
        col_letter = get_column_letter(col)
        new_record[header] = validate_field(header, col_letter)
        sheet.cell(row=first_empty_row, column=col).value = new_record[header]

    print("\nВы заполнили следующие поля:")
    for header, value in new_record.items():
        print(f"{header}: {value}")

    while True:
        confirm = input("Сохранить изменения? (1 - 'Да', 2 - 'Редактировать'): ").strip()
        if confirm == "1":
            workbook.save(file_name)
            print("Запись успешно добавлена и сохранена.")
            break
        elif confirm == "2":
            edit_record_logic(sheet, first_empty_row, new_record)
            workbook.save(file_name)
            print("Изменения сохранены после редактирования.")
            break
        else:
            print("Некорректный ввод. Введите 1 или 2.")

# Редактирование записи
def edit_record_logic(sheet, row, record):
    print("\nРедактирование записи:")
    for col, (header, value) in enumerate(record.items(), start=1):
        print(f"{col}. {header}: {value}")

    while True:
        try:
            field_num = int(input("Введите номер поля для редактирования (или 0 для завершения): ").strip())
            if field_num == 0:
                break
            if 1 <= field_num <= len(record):
                header = list(record.keys())[field_num - 1]
                col_letter = get_column_letter(field_num)
                current_value = record.get(header)

                new_value = validate_field(header, col_letter, current_value)
                if new_value is not None:
                    record[header] = new_value
                    sheet.cell(row=row, column=field_num).value = new_value
                    print(f"Поле '{header}' изменено.")
            else:
                print("Некорректный номер поля. Попробуйте снова.")
        except ValueError:
            print("Введите корректное число.")

# Настройка таблицы
def configure_table(file_name):
    print("Настройка таблицы пока не реализована.")

# Найти первую пустую строку
def find_first_empty_row(sheet):
    for row in range(2, sheet.max_row + 2):
        if not any(sheet.cell(row=row, column=col).value for col in range(1, sheet.max_column + 1)):
            return row

if __name__ == "__main__":
    main()
