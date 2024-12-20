import os
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
import re


def find_file_in_yandex_disk(file_name):
    """Ищет файл в папке Яндекс.Диска."""
    # Предполагаем, что Яндекс.Диск синхронизирован в стандартной папке
    yandex_disk_path = os.path.expanduser(r"~\\YandexDisk")
    for root, dirs, files in os.walk(yandex_disk_path):
        if file_name in files:
            return os.path.join(root, file_name)
    return None

def main():
    # Устанавливаем текущую рабочую директорию на директорию скрипта/исполняемого файла
    if getattr(sys, 'frozen', False):  # Если это .exe файл
        os.chdir(os.path.dirname(sys.executable))
    else:  # Если это обычный .py скрипт
        os.chdir(os.path.dirname(__file__))

    print(f"Текущая директория установлена: {os.getcwd()}")

    # Проверяем наличие файла в текущей папке или на Яндекс.Диске
    file_name = "_РЕЕСТР МАТЕРИАЛОВ.xlsx"
    current_directory = os.getcwd()
    file_path = os.path.join(current_directory, file_name)

    if not os.path.exists(file_path):
        print(f"Файл {file_name} не найден в текущей директории.")
        print("Ищу файл в папке Яндекс.Диска...")
        file_path = find_file_in_yandex_disk(file_name)

        if not file_path:
            print("Файл не найден. Введите абсолютный путь к файлу:")
            while True:
                file_path = input("Введите путь к файлу: ").strip()
                if os.path.exists(file_path) and file_path.endswith((".xlsx", ".xlsm")):
                    break
                print("Файл не найден или это не Excel-файл. Попробуйте снова.")
        else:
            print(f"Файл найден на Яндекс.Диске: {file_path}")
    else:
        print(f"Файл {file_name} обнаружен в текущей директории.")

    # Проверка ввода ключевого слова "реестр"
    while True:
        word = input("Введите слово 'реестр': ").strip().lower()
        if word == "реестр":
            break
        print("Нет, именно 'реестр':)")

    current_directory = os.getcwd()
    print(f"Текущая директория: {current_directory}")

    # Главное меню
    while True:
        print("\nГлавное меню:")
        print("1. Открыть реестр в Excel")
        print("2. Добавить запись")
        print("3. Изменить запись")
        print("4. Настроить таблицу")
        print("5. Выход")

        choice = input("Выберите пункт меню: ").strip()

        if choice == "1":
            open_registry(file_name)
        elif choice == "2":
            add_record(file_name)
        elif choice == "3":
            edit_record(file_name)
        elif choice == "4":
            configure_table(file_name)
        elif choice == "5":
            print("Выход из программы. До свидания!")
            break
        else:
            print("Некорректный ввод. Выберите пункт от 1 до 5.")

# Открыть реестр в Excel
def open_registry(file_name):
    try:
        print(f"Открываю файл {file_name}...")
        os.startfile(file_name)
    except Exception as e:
        print(f"Не удалось открыть файл. Ошибка: {e}")

# Добавить запись
def add_record(file_name):
    try:
        workbook = load_workbook(file_name)
        sheet = workbook.active
    except Exception as e:
        print(f"Ошибка при открытии файла: {e}")
        return

    print("\nДобавление новой записи")
    first_empty_row = find_first_empty_row(sheet)

    # Заголовки столбцов
    headers = [sheet.cell(row=1, column=col).value for col in range(1, sheet.max_column + 1)]
    section_choices = [
        "Банк практик", "Видеолекции", "Видеоматериалы", "Жизненные ситуации",
        "Инструкция по внедрению", "Исследования и обзоры", "Публикации",
        "Отечественный опыт", "Международный опыт", "Иное"
    ]
    new_record = {}

    for col, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col)
        if col_letter in {"A", "B"}:
            continue  # Пропускаем A и B для обработки позже

        if not header:
            continue

        while True:
            if col_letter == "C" or col_letter == "D":
                value = input(f"Введите значение для {header}: ").strip()
                if value:
                    break
                print(f"Поле {header} обязательно для заполнения. Попробуйте снова.")
            elif col_letter == "E":
                value = input(f"Введите значение для аннотации (макс. 250 символов) (на обложку материала): ").strip()
                if value and len(value) <= 250:
                    break
                if not value:
                    print("Поле обязательно для заполнения. Попробуйте снова.")
                else:
                    print("Аннотация не может быть длиннее 250 символов. Попробуйте снова.")
            elif col_letter == "F":
                print("Выберите раздел:")
                for i, choice in enumerate(section_choices, start=1):
                    print(f"{i}. {choice}")
                choice = input("Введите номер раздела: ").strip()
                if choice.isdigit() and 1 <= int(choice) <= len(section_choices):
                    value = section_choices[int(choice) - 1]
                    if value == "Иное":
                        value = input("Введите название для 'Иное': ").strip()
                    break
                print("Некорректный выбор. Введите номер из списка.")
            elif col_letter == "G":
                value = input(f"Введите значение для {header} (или оставьте пустым): ").strip()
                break
            elif col_letter == "H" or col_letter == "I":
                value = input(f"Введите значение для {header} (в формате ДД-ММ-ГГГГ или оставьте пустым): ").strip()
                if not value:  # Поле может быть пустым
                    break
                if re.match(r"^\d{2}-\d{2}-\d{4}$", value):
                    break
                print(f"Некорректный формат даты. Введите дату в формате ДД-ММ-ГГГГ или оставьте пустым.")
            elif col_letter == "J":
                value = input(f"{header} (1 - 'Нет', 2 - 'Да'): ").strip()
                if value == "1":
                    print("Тогда разместите файл и потом выберите 'Да'.")
                    continue
                elif value == "2":
                    value = "Да"
                else:
                    print("Некорректный ввод. Введите 1 или 2.")
                    continue
                break
            elif col_letter == "K":
                value = input(f"Введите значение для {header}: ").strip()
                if "/" in value:
                    break
                print("Ссылка должна содержать символ '/'. Попробуйте снова.")
            elif col_letter == "L":
                value = input(f"{header} (1 - 'Текст', 2 - 'Видео'): ").strip()
                if value == "1":
                    value = "Текст"
                elif value == "2":
                    value = "Видео"
                    print("Формат 'Видео' выбран. Столбцы 'Код внедрения' и 'Код внедрения (адаптированный)' пропускаются.")
                else:
                    print("Некорректный ввод. Введите 1 или 2.")
                    continue
                new_record["Формат"] = value
                break
            elif col_letter == "M":
                if new_record.get("Формат", "") == "Видео":
                    break
                value = input(f"{header}: ").strip()
                if "<iframe src=" in value:
                    break
                print("Введите корректный код внедрения OneDrive.")
            elif col_letter == "N":
                if new_record.get("Формат", "") == "Видео":
                    break
                iframe_code = new_record.get("Код внедрения (OneDrive, если текст)", "")
                if iframe_code:
                    # Обновление width и height в iframe
                    adapted_code = re.sub(r'width="[^"]*"', 'width="90%"', iframe_code)
                    adapted_code = re.sub(r'height="[^"]*"', 'height="1800"', adapted_code)
                    value = f'<p align="center">{adapted_code}</p>'
                    print(f"Код внедрения адаптирован для отображения на tilda: {value}")
                else:
                    value = None
                break

            elif col_letter == "O":
                value = input(f"Введите имя страницы Tilda (латиница): ").strip()
                if value.isascii() and value.isalnum():
                    break
                print("Поле может содержать только латинские буквы и цифры. Попробуйте снова.")
            elif col_letter == "P":
                value = input(f"{header} (или оставьте пустым): ").strip()
                break
            else:
                value = input(f"Введите значение для {header}: ").strip()
                if value:
                    break
                print(f"Поле {header} обязательно для заполнения. Попробуйте снова.")

        new_record[header] = value
        sheet.cell(row=first_empty_row, column=col).value = value

    # Обработка столбцов A и B
    for col_letter, header in [("A", headers[0]), ("B", headers[1])]:
        while True:
            value = input(f"{header} (1 - 'Нет', 2 - 'Да'): ").strip()
            if value == "1":
                value = "Нет"
            elif value == "2":
                value = "Да"
            else:
                print("Некорректный ввод. Введите 1 или 2.")
                continue
            break
        new_record[header] = value
        sheet.cell(row=first_empty_row, column=ord(col_letter) - 64).value = value

    # Подтверждение данных перед сохранением
    print("\nВы заполнили следующие поля:")
    for header, value in new_record.items():
        print(f"{header}: {value}")

    while True:
        confirm = input("Сохранить изменения? (1 - 'Да', 2 - 'Редактировать'): ").strip()
        if confirm == "1":
            workbook.save(file_name)
            print("Запись успешно добавлена и сохранена.")
            break
        elif confirm == "2":
            edit_record_logic(sheet, first_empty_row, new_record)
            workbook.save(file_name)
            print("Изменения сохранены после редактирования.")
            break
        else:
            print("Некорректный ввод. Введите 1 или 2.")

# Логика редактирования записи
def edit_record_logic(sheet, row, record):
    print("\nРедактирование записи:")
    for col, (header, value) in enumerate(record.items(), start=1):
        print(f"{col}. {header}: {value}")

    while True:
        try:
            field_num = int(input("Введите номер поля для редактирования (или 0 для завершения): ").strip())
            if field_num == 0:
                break
            if 1 <= field_num <= len(record):
                header = list(record.keys())[field_num - 1]
                col_letter = get_column_letter(field_num)

                # Логика для редактирования
                while True:
                    if col_letter == "A" or col_letter == "B":
                        new_value = input(f"{header} (1 - 'Нет', 2 - 'Да'): ").strip()
                        if new_value == "1":
                            new_value = "Нет"
                        elif new_value == "2":
                            new_value = "Да"
                        else:
                            print("Некорректный ввод. Введите 1 или 2.")
                            continue
                        break
                    elif col_letter == "C" or col_letter == "D":
                        new_value = input(f"Введите значение для {header}: ").strip()
                        if new_value:
                            break
                        print(f"Поле {header} обязательно для заполнения. Попробуйте снова.")
                    elif col_letter == "E":
                        new_value = input(f"Введите значение для аннотации (макс. 250 символов) (на обложку материала): ").strip()
                        if len(new_value) <= 250:
                            break
                        print("Аннотация не может быть длиннее 250 символов. Попробуйте снова.")
                    elif col_letter == "F":
                        section_choices = [
                            "Банк практик", "Видеолекции", "Видеоматериалы", "Жизненные ситуации",
                            "Инструкция по внедрению", "Исследования и обзоры", "Публикации",
                            "Отечественный опыт", "Международный опыт", "Иное"
                        ]
                        print("Выберите раздел:")
                        for i, choice in enumerate(section_choices, start=1):
                            print(f"{i}. {choice}")
                        choice = input("Введите номер раздела: ").strip()
                        if choice.isdigit() and 1 <= int(choice) <= len(section_choices):
                            new_value = section_choices[int(choice) - 1]
                            if new_value == "Иное":
                                new_value = input("Введите название для 'Иное': ").strip()
                            break
                        print("Некорректный выбор. Введите номер из списка.")
                    elif col_letter == "H" or col_letter == "I":
                        new_value = input(f"Введите значение для {header} (в формате ДД-ММ-ГГГГ или оставьте пустым): ").strip()
                        if not new_value or re.match(r"^\\d{2}-\\d{2}-\\d{4}$", new_value):
                            break
                        print("Некорректный формат даты. Введите дату в формате ДД-ММ-ГГГГ или оставьте пустым.")
                    elif col_letter == "K":
                        new_value = input(f"Введите значение для {header}: ").strip()
                        if "/" in new_value:
                            break
                        print("Ссылка должна содержать символ '/'. Попробуйте снова.")
                    elif col_letter == "L":
                        new_value = input(f"{header} (1 - 'Текст', 2 - 'Видео'): ").strip()
                        if new_value == "1":
                            new_value = "Текст"
                        elif new_value == "2":
                            new_value = "Видео"
                            print("Формат 'Видео' выбран. Поля 'Код внедрения' и 'Код внедрения (адаптированный)' пропускаются.")
                        else:
                            print("Некорректный ввод. Введите 1 или 2.")
                            continue
                        break
                    elif col_letter == "M":
                        if record.get("Формат", "") == "Видео":
                            print("Формат 'Видео'. Поле пропускается.")
                            break
                        new_value = input(f"{header}: ").strip()
                        if "<iframe src=" in new_value:
                            break
                        print("Введите корректный код внедрения OneDrive.")
                    elif col_letter == "N":
                        if record.get("Формат", "") == "Видео":
                            print("Формат 'Видео'. Поле пропускается.")
                            break
                        iframe_code = record.get("Код внедрения (OneDrive, если текст)", "")
                        if iframe_code:
                            adapted_code = re.sub(r'width="[^"]*"', 'width="90%"', iframe_code)
                            adapted_code = re.sub(r'height="[^"]*"', 'height="1800"', iframe_code)
                            new_value = f'<p align="center">{adapted_code}</p>'
                            print(f"Код внедрения адаптирован для отображения на tilda: {new_value}")
                        else:
                            new_value = None
                        break
                    elif col_letter == "O":
                        new_value = input(f"Введите имя страницы Tilda (латиница): ").strip()
                        if new_value.isascii() and new_value.isalnum():
                            break
                        print("Поле может содержать только латинские буквы и цифры. Попробуйте снова.")
                    elif col_letter == "P":
                        new_value = input(f"{header} (или оставьте пустым): ").strip()
                        break
                    else:
                        new_value = input(f"Введите значение для {header}: ").strip()
                        if new_value:
                            break
                        print(f"Поле {header} обязательно для заполнения. Попробуйте снова.")

                # Сохранение нового значения
                record[header] = new_value
                sheet.cell(row=row, column=field_num).value = new_value
                print(f"Поле '{header}' изменено.")
            else:
                print("Некорректный номер поля. Попробуйте снова.")
        except ValueError:
            print("Введите корректное число.")




# Настройка таблицы
def configure_table(file_name):
    print("Настройка таблицы пока не реализована.")

# Найти первую пустую строку
def find_first_empty_row(sheet):
    for row in range(2, sheet.max_row + 2):
        if not any(sheet.cell(row=row, column=col).value for col in range(1, sheet.max_column + 1)):
            return row

if __name__ == "__main__":
    main()
